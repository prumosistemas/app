import asyncio
import base64
import hashlib
import os
import re
import secrets
import shutil
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import Header, HTTPException, Request
from cryptography.fernet import Fernet, InvalidToken
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from flow_core import somente_digitos

from db import (
    ACTIVE_STATUSES,
    ALLOW_DIRECT_LOCAL,
    BASE_BROWSER_SLOTS,
    BROWSER_POOL_CONFIGURED,
    BROWSER_TURBO_EXTRA,
    DB_FILE,
    FINAL_STATUSES,
    FLOW_EXECUTION_ORDER,
    FLOW_LABELS,
    FLOW_ORDER,
    HEADLESS,
    ISS_INTERNAL_SECRET,
    ITEM_FINAL_STATUSES,
    MAX_BROWSERS,
    MAX_DATASETS,
    MAX_RUNS_PER_MEMBER,
    OUTPUT_ROOT,
    RUN_RETENTION_DAYS,
    DATA_ROOT,
    WORKER_PUBLIC_URL,
    db_get_json,
    db_set_json,
    now_ms,
)


RUNS: Dict[str, Dict[str, Any]] = {}
RUN_LOCK = asyncio.Lock()
SECRET_ENC_PREFIX = "enc:v1:"
_FERNET_CACHE: Optional[Fernet] = None


@dataclass(frozen=True)
class WorkerContext:
    company_id: str
    company_name: str
    user_id: str
    user_email: str
    user_role: str
    via_worker: bool


class AccountPayload(BaseModel):
    alias: str
    usuario: str
    senha: str


class DatasetItem(BaseModel):
    cnpj: str
    codigo_dominio: str = ""
    nome_empresa: str = ""
    account_id: str = ""


class DatasetPayload(BaseModel):
    alias: str
    items: List[DatasetItem]


class RunCreateRequest(BaseModel):
    mes_num: int = Field(..., description="Mês numérico, 1 a 12")
    ano: int = Field(..., description="Ano com 4 dígitos")
    dataset_id: str
    flow_selection: Dict[str, Dict[str, bool]] = Field(default_factory=dict)
    usar_codigo_dominio: bool = True
    reabrir_escrituracao_fechada: bool = True


class RetryRequest(BaseModel):
    only_retryable: bool = False
    include_cancelled: bool = True
    include_interrupted: bool = False


class StopFlowRequest(BaseModel):
    cnpj: str
    flow_mode: str


def safe_slug(value: str, fallback: str = "sem-id") -> str:
    value = str(value or "").strip()
    value = re.sub(r"[^A-Za-z0-9_.-]+", "_", value)
    value = value.strip("._-")
    return value[:90] or fallback


def secret_fernet() -> Optional[Fernet]:
    global _FERNET_CACHE

    if _FERNET_CACHE is not None:
        return _FERNET_CACHE

    if not ISS_INTERNAL_SECRET:
        return None

    digest = hashlib.sha256(("prumo-secret-fields-v1:" + ISS_INTERNAL_SECRET).encode("utf-8")).digest()
    key = base64.urlsafe_b64encode(digest)
    _FERNET_CACHE = Fernet(key)
    return _FERNET_CACHE


def protect_secret(value: Any) -> str:
    raw = str(value or "")

    if not raw or raw.startswith(SECRET_ENC_PREFIX):
        return raw

    fernet = secret_fernet()
    if fernet is None:
        return raw

    token = fernet.encrypt(raw.encode("utf-8")).decode("ascii")
    return SECRET_ENC_PREFIX + token


def unprotect_secret(value: Any) -> str:
    raw = str(value or "")

    if not raw.startswith(SECRET_ENC_PREFIX):
        return raw

    fernet = secret_fernet()
    if fernet is None:
        return ""

    token = raw[len(SECRET_ENC_PREFIX):]
    try:
        return fernet.decrypt(token.encode("ascii")).decode("utf-8")
    except (InvalidToken, UnicodeDecodeError, ValueError):
        return ""


def protect_account_for_storage(acc: Dict[str, Any]) -> Dict[str, Any]:
    stored = dict(acc)
    stored["usuario"] = protect_secret(stored.get("usuario", ""))
    stored["senha"] = protect_secret(stored.get("senha", ""))
    return stored


def unprotect_account_from_storage(acc: Dict[str, Any]) -> Dict[str, Any]:
    loaded = dict(acc)
    loaded["usuario"] = unprotect_secret(loaded.get("usuario", ""))
    loaded["senha"] = unprotect_secret(loaded.get("senha", ""))
    return loaded


def protect_credential_snapshots_for_storage(snapshots: Dict[str, Any]) -> Dict[str, Any]:
    output: Dict[str, Any] = {}

    for key, value in (snapshots or {}).items():
        snap = dict(value or {})
        snap["usuario"] = protect_secret(snap.get("usuario", ""))
        snap["senha"] = protect_secret(snap.get("senha", ""))
        output[key] = snap

    return output


def unprotect_credential_snapshots_from_storage(snapshots: Dict[str, Any]) -> Dict[str, Any]:
    output: Dict[str, Any] = {}

    for key, value in (snapshots or {}).items():
        snap = dict(value or {})
        snap["usuario"] = unprotect_secret(snap.get("usuario", ""))
        snap["senha"] = unprotect_secret(snap.get("senha", ""))
        output[key] = snap

    return output


def direct_local_context() -> WorkerContext:
    return WorkerContext(
        company_id="local_dev_empresa",
        company_name="Empresa Local DEV",
        user_id="local_dev_usuario",
        user_email="local@dev",
        user_role="master",
        via_worker=False,
    )


async def get_worker_context(
    request: Request,
    x_internal_secret: str = Header(default="", alias="X-Internal-Secret"),
    x_company_id: str = Header(default="", alias="X-Company-Id"),
    x_company_name: str = Header(default="", alias="X-Company-Name"),
    x_user_id: str = Header(default="", alias="X-User-Id"),
    x_user_email: str = Header(default="", alias="X-User-Email"),
    x_user_role: str = Header(default="", alias="X-User-Role"),
) -> WorkerContext:
    has_worker_headers = bool(x_company_id and x_user_id and x_user_email)

    if has_worker_headers:
        if not ISS_INTERNAL_SECRET:
            raise HTTPException(status_code=500, detail="ISS_INTERNAL_SECRET não configurado.")

        if x_internal_secret != ISS_INTERNAL_SECRET:
            raise HTTPException(status_code=403, detail="Segredo interno inválido.")

        return WorkerContext(
            company_id=safe_slug(x_company_id),
            company_name=str(x_company_name or x_company_id),
            user_id=safe_slug(x_user_id),
            user_email=str(x_user_email),
            user_role=str(x_user_role or "member"),
            via_worker=True,
        )

    client_host = request.client.host if request.client else ""
    is_local_host = client_host in {"127.0.0.1", "::1", "localhost"}

    if ALLOW_DIRECT_LOCAL and is_local_host:
        return direct_local_context()

    raise HTTPException(status_code=401, detail="Acesso permitido somente via Worker.")


def scope_id(ctx: WorkerContext) -> str:
    return f"{safe_slug(ctx.company_id)}:{safe_slug(ctx.user_id)}"


def owner_key(ctx: WorkerContext) -> str:
    return scope_id(ctx)


def member_prefix(ctx: WorkerContext) -> str:
    return f"empresa:{safe_slug(ctx.company_id)}:membro:{safe_slug(ctx.user_id)}"


def kv_key(ctx: WorkerContext, name: str) -> str:
    return f"{member_prefix(ctx)}:{name}"


def company_output_root(ctx: WorkerContext) -> str:
    path = os.path.join(OUTPUT_ROOT, "empresas", safe_slug(ctx.company_id))
    Path(path).mkdir(parents=True, exist_ok=True)
    return path


def member_output_root(ctx: WorkerContext) -> str:
    path = os.path.join(company_output_root(ctx), "colaboradores", safe_slug(ctx.user_id))
    Path(path).mkdir(parents=True, exist_ok=True)
    return path


def member_runs_root(ctx: WorkerContext) -> str:
    path = os.path.join(member_output_root(ctx), "runs")
    Path(path).mkdir(parents=True, exist_ok=True)
    return path


def company_user_root(ctx: WorkerContext) -> str:
    return member_output_root(ctx)


def new_id(prefix: str) -> str:
    return f"{prefix}_{secrets.token_urlsafe(16)}"


def new_run_id() -> str:
    return new_id("run")


def new_dataset_id() -> str:
    return new_id("set")


def new_account_id() -> str:
    return new_id("acc")


def model_to_dict(item: Any) -> Dict[str, Any]:
    if hasattr(item, "model_dump"):
        return item.model_dump()
    if hasattr(item, "dict"):
        return item.dict()
    return dict(item)


def clean_alias(alias: str) -> str:
    alias = str(alias or "").strip()
    alias = re.sub(r"\s+", " ", alias)
    return alias[:90]


def only_digits(value: str) -> str:
    return somente_digitos(str(value or ""))


def normalize_cnpj(cnpj: str) -> str:
    return only_digits(cnpj).zfill(14)


def format_cnpj(cnpj: str) -> str:
    d = normalize_cnpj(cnpj)
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


def valid_cnpj_basic(cnpj: str) -> bool:
    d = only_digits(cnpj)
    if len(d) != 14:
        return False
    if len(set(d)) == 1:
        return False
    return True


def build_mes(mes_num: int, ano: int) -> str:
    try:
        mes_num_int = int(mes_num)
        ano_int = int(ano)
    except Exception:
        raise HTTPException(status_code=400, detail="Mês e ano inválidos.")

    if mes_num_int < 1 or mes_num_int > 12:
        raise HTTPException(status_code=400, detail="Mês inválido.")

    if ano_int < 2000 or ano_int > 2100:
        raise HTTPException(status_code=400, detail="Ano inválido.")

    return f"{mes_num_int:02d}/{ano_int}"


def default_flows() -> Dict[str, bool]:
    return {flow: True for flow in FLOW_ORDER}


def normalize_flows(flows: Optional[Dict[str, Any]]) -> Dict[str, bool]:
    if not isinstance(flows, dict):
        return default_flows()
    return {flow: bool(flows.get(flow, False)) for flow in FLOW_ORDER}


def safe_path_inside(base: str, target: str) -> str:
    base_real = os.path.realpath(base)
    target_real = os.path.realpath(target)

    try:
        common = os.path.commonpath([base_real, target_real])
    except ValueError:
        raise HTTPException(status_code=400, detail="Caminho inválido.")

    if common != base_real:
        raise HTTPException(status_code=400, detail="Caminho inválido.")

    return target_real


def task_key(item_or_result: Dict[str, Any]) -> str:
    return f"{normalize_cnpj(item_or_result.get('cnpj', ''))}|{item_or_result.get('flow_mode', '')}"


def result_merge_key(result: Dict[str, Any]) -> str:
    return task_key(result)


def compact_account(acc: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": acc.get("id"),
        "alias": acc.get("alias"),
        "usuario": acc.get("usuario"),
        "created_at": acc.get("created_at"),
        "updated_at": acc.get("updated_at"),
    }


def strip_runtime_task(item: Dict[str, Any]) -> Dict[str, Any]:
    flow_mode = item.get("flow_mode", "")
    return {
        "cnpj": item.get("cnpj", ""),
        "cnpj_digits": item.get("cnpj_digits", ""),
        "codigo_dominio": item.get("codigo_dominio", ""),
        "nome_empresa": item.get("nome_empresa", ""),
        "account_id": item.get("account_id", ""),
        "account_alias": item.get("account_alias", ""),
        "flow_mode": flow_mode,
        "flow_label": FLOW_LABELS.get(flow_mode, flow_mode),
        "usar_codigo_dominio": bool(item.get("usar_codigo_dominio", True)),
        "reabrir_escrituracao_fechada": bool(item.get("reabrir_escrituracao_fechada", True)),
    }


def strip_runtime_tasks(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [strip_runtime_task(item) for item in items]


def credential_snapshot_from_tasks(runtime_tasks: List[Dict[str, Any]]) -> Dict[str, Dict[str, str]]:
    snapshots: Dict[str, Dict[str, str]] = {}
    for item in runtime_tasks:
        snapshots[task_key(item)] = {
            "account_id": str(item.get("account_id", "")),
            "account_alias": str(item.get("account_alias", "")),
            "usuario": str(item.get("usuario", "")),
            "senha": str(item.get("senha", "")),
        }
    return snapshots


def sanitize_run_for_response(run: Dict[str, Any]) -> Dict[str, Any]:
    safe = dict(run)
    safe.pop("credential_snapshots", None)
    return safe


def load_accounts_raw(ctx: WorkerContext) -> List[Dict[str, Any]]:
    payload = db_get_json(kv_key(ctx, "accounts"), {"accounts": []})
    accounts = payload.get("accounts", [])
    if not isinstance(accounts, list):
        return []
    return [unprotect_account_from_storage(acc) for acc in accounts if isinstance(acc, dict)]


def save_accounts(ctx: WorkerContext, accounts: List[Dict[str, Any]]) -> None:
    db_set_json(
        kv_key(ctx, "accounts"),
        {
            "updated_at": now_ms(),
            "accounts": [protect_account_for_storage(acc) for acc in accounts],
        },
    )


def load_accounts_public(ctx: WorkerContext) -> List[Dict[str, Any]]:
    return [compact_account(acc) for acc in load_accounts_raw(ctx)]


def get_account_or_404(ctx: WorkerContext, account_id: str) -> Dict[str, Any]:
    for acc in load_accounts_raw(ctx):
        if acc.get("id") == account_id:
            return acc
    raise HTTPException(status_code=404, detail="Conta não encontrada.")


def account_ref_to_id(ctx: WorkerContext, ref: str) -> str:
    ref = str(ref or "").strip()
    if not ref:
        return ""

    accounts = load_accounts_raw(ctx)

    for acc in accounts:
        if acc.get("id") == ref:
            return acc["id"]

    ref_low = ref.lower()
    for acc in accounts:
        if str(acc.get("alias", "")).lower() == ref_low:
            return acc["id"]

    return ref


def account_exists(ctx: WorkerContext, account_id: str) -> bool:
    return bool(account_id) and any(acc.get("id") == account_id for acc in load_accounts_raw(ctx))


def load_datasets(ctx: WorkerContext) -> List[Dict[str, Any]]:
    payload = db_get_json(kv_key(ctx, "datasets"), {"datasets": []})
    datasets = payload.get("datasets", [])
    return datasets if isinstance(datasets, list) else []


def compact_dataset_record(dataset: Dict[str, Any]) -> Dict[str, Any]:
    items = dataset.get("items", []) or []
    data = {key: value for key, value in dataset.items() if key != "items"}
    data["count"] = dataset.get("count", len(items))
    return data


def save_datasets(ctx: WorkerContext, datasets: List[Dict[str, Any]]) -> None:
    db_set_json(
        kv_key(ctx, "datasets"),
        {
            "updated_at": now_ms(),
            "max_datasets": MAX_DATASETS if MAX_DATASETS > 0 else None,
            "datasets": datasets[:MAX_DATASETS] if MAX_DATASETS > 0 else datasets,
        },
    )


def get_dataset_or_404(ctx: WorkerContext, dataset_id: str) -> Dict[str, Any]:
    for ds in load_datasets(ctx):
        if ds.get("id") == dataset_id:
            return ds
    raise HTTPException(status_code=404, detail="Conjunto não encontrado.")


def account_usage(ctx: WorkerContext) -> Dict[str, List[Dict[str, str]]]:
    usage: Dict[str, List[Dict[str, str]]] = {}

    for ds in load_datasets(ctx):
        for item in ds.get("items", []):
            account_id = item.get("account_id", "")
            if not account_id:
                continue

            usage.setdefault(account_id, []).append(
                {
                    "dataset_id": ds.get("id", ""),
                    "dataset_alias": ds.get("alias", ""),
                    "cnpj": item.get("cnpj", ""),
                    "nome_empresa": item.get("nome_empresa", ""),
                    "codigo_dominio": item.get("codigo_dominio", ""),
                }
            )

    return usage


def normalize_dataset_item(raw: Any) -> Dict[str, Any]:
    if isinstance(raw, DatasetItem):
        return {
            "cnpj": str(raw.cnpj or "").strip(),
            "codigo_dominio": str(raw.codigo_dominio or "").strip(),
            "nome_empresa": str(raw.nome_empresa or "").strip(),
            "account_id": str(raw.account_id or "").strip(),
        }

    return {
        "cnpj": str(raw.get("cnpj", "")).strip(),
        "codigo_dominio": str(raw.get("codigo_dominio", "")).strip(),
        "nome_empresa": str(raw.get("nome_empresa", "")).strip(),
        "account_id": str(raw.get("account_id", "") or raw.get("conta", "") or raw.get("account_ref", "")).strip(),
    }


def validate_dataset_items(ctx: WorkerContext, items: List[Any]) -> Dict[str, Any]:
    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen: Dict[str, int] = {}

    for idx, raw in enumerate(items, start=1):
        item = normalize_dataset_item(raw)

        cnpj_raw = item["cnpj"]
        cnpj_digits = only_digits(cnpj_raw)
        codigo = item["codigo_dominio"]
        nome_empresa = item["nome_empresa"]
        account_id = account_ref_to_id(ctx, item["account_id"])
        row_errors: List[str] = []

        if not valid_cnpj_basic(cnpj_raw):
            row_errors.append("CNPJ inválido. Informe 14 dígitos válidos.")

        if valid_cnpj_basic(cnpj_raw):
            cnpj_norm = cnpj_digits.zfill(14)
            if cnpj_norm in seen:
                row_errors.append(f"CNPJ duplicado. Já apareceu na linha {seen[cnpj_norm]}.")
            else:
                seen[cnpj_norm] = idx
        else:
            cnpj_norm = cnpj_digits

        if codigo and not re.fullmatch(r"[A-Za-z0-9._-]+", codigo):
            row_errors.append("Código Domínio contém caracteres inválidos.")

        if not account_id:
            row_errors.append("Conta obrigatória para este CNPJ.")
        elif not account_exists(ctx, account_id):
            row_errors.append("Conta não encontrada. Use uma conta cadastrada.")

        row = {
            "linha": idx,
            "cnpj": format_cnpj(cnpj_norm) if len(cnpj_norm) == 14 else cnpj_raw,
            "cnpj_digits": cnpj_norm,
            "codigo_dominio": codigo,
            "nome_empresa": nome_empresa,
            "account_id": account_id,
            "valid": len(row_errors) == 0,
            "errors": row_errors,
        }

        rows.append(row)

        if row_errors:
            errors.append(row)

    return {
        "valid": len(errors) == 0 and len(rows) > 0,
        "total": len(rows),
        "validos": len([r for r in rows if r["valid"]]),
        "invalidos": len(errors),
        "items": rows,
        "errors": errors,
    }


def normalize_valid_dataset_items(validated_items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    output: List[Dict[str, Any]] = []

    for item in validated_items:
        if not item.get("valid"):
            continue

        output.append(
            {
                "cnpj": item["cnpj"],
                "cnpj_digits": item["cnpj_digits"],
                "codigo_dominio": item.get("codigo_dominio", ""),
                "nome_empresa": item.get("nome_empresa", ""),
                "account_id": item.get("account_id", ""),
            }
        )

    return output


def save_dataset_record(ctx: WorkerContext, dataset_id: Optional[str], alias: str, items: List[Any]) -> Dict[str, Any]:
    alias = clean_alias(alias)

    if not alias:
        raise HTTPException(status_code=400, detail="Informe o alias do conjunto.")

    validation = validate_dataset_items(ctx, items)

    if not validation["valid"]:
        return {
            "saved": False,
            "message": "Conjunto possui erros.",
            "validation": validation,
            "datasets": load_datasets(ctx),
        }

    datasets = load_datasets(ctx)
    now = now_ms()
    normalized_items = normalize_valid_dataset_items(validation["items"])

    if dataset_id:
        found = False
        saved = None

        for ds in datasets:
            if ds.get("id") == dataset_id:
                ds.update(
                    {
                        "alias": alias,
                        "items": normalized_items,
                        "count": len(normalized_items),
                        "updated_at": now,
                    }
                )
                found = True
                saved = ds
                break

        if not found:
            raise HTTPException(status_code=404, detail="Conjunto não encontrado.")
    else:
        if MAX_DATASETS > 0 and len(datasets) >= MAX_DATASETS:
            raise HTTPException(status_code=400, detail=f"Limite de {MAX_DATASETS} conjuntos atingido.")

        saved = {
            "id": new_dataset_id(),
            "alias": alias,
            "items": normalized_items,
            "count": len(normalized_items),
            "created_at": now,
            "updated_at": now,
        }

        datasets.insert(0, saved)

    save_datasets(ctx, datasets)

    return {
        "saved": True,
        "dataset": saved,
        "datasets": datasets,
        "validation": validation,
    }


def read_xlsx_items_from_bytes(ctx: WorkerContext, data: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    items: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cnpj = row[0] if len(row) >= 1 else None
        codigo = row[1] if len(row) >= 2 else None
        nome = row[2] if len(row) >= 3 else None
        conta = row[3] if len(row) >= 4 else None

        if cnpj is None and codigo is None and nome is None and conta is None:
            continue

        cnpj_str = str(cnpj or "").strip()
        codigo_str = str(codigo or "").strip()
        nome_str = str(nome or "").strip()
        conta_str = str(conta or "").strip()

        if row_idx == 1:
            header = " ".join([cnpj_str, codigo_str, nome_str, conta_str]).lower()
            if any(token in header for token in ["cnpj", "dominio", "domínio", "empresa", "conta", "login"]):
                continue

        items.append(
            {
                "cnpj": cnpj_str,
                "codigo_dominio": codigo_str,
                "nome_empresa": nome_str,
                "account_id": account_ref_to_id(ctx, conta_str),
            }
        )

    wb.close()
    return items


def load_runs_state(ctx: WorkerContext) -> Dict[str, Dict[str, Any]]:
    payload = db_get_json(kv_key(ctx, "runs_state"), {"runs": {}})
    runs = payload.get("runs", {})

    if not isinstance(runs, dict):
        return {}

    for run in runs.values():
        run["credential_snapshots"] = unprotect_credential_snapshots_from_storage(
            run.get("credential_snapshots", {}) or {}
        )

        if run.get("status") in ACTIVE_STATUSES:
            run["status"] = "interrompida"
            run["finished_at"] = run.get("finished_at") or now_ms()
            updated_results = []
            for result in run.get("results", []) or []:
                if result.get("status") == "running":
                    updated = dict(result)
                    updated["status"] = "interrompida"
                    updated["erro"] = "Execução interrompida antes de finalizar."
                    updated["erro_code"] = "RUN_INTERRUPTED"
                    updated["erro_action"] = "Executar retry se este fluxo ainda for necessário."
                    updated["retryable"] = True
                    updated["finished_at"] = updated.get("finished_at") or now_ms()
                    updated_results.append(updated)
                else:
                    updated_results.append(result)
            run["results"] = updated_results

    return runs


def save_runs_state(ctx: WorkerContext) -> None:
    _prune_run_retention_in_memory(ctx)

    safe: Dict[str, Dict[str, Any]] = {}
    prefix = f"{scope_id(ctx)}:"

    for run_key, run in RUNS.items():
        if run.get("scope_id") != scope_id(ctx):
            continue

        item = dict(run)
        item.pop("logs", None)
        item["credential_snapshots"] = protect_credential_snapshots_for_storage(
            item.get("credential_snapshots", {}) or {}
        )
        safe[run_key.replace(prefix, "", 1)] = item

    db_set_json(kv_key(ctx, "runs_state"), {"updated_at": now_ms(), "runs": safe})


def ensure_member_runs_loaded(ctx: WorkerContext) -> None:
    prefix = f"{scope_id(ctx)}:"

    already = any(r.get("scope_id") == scope_id(ctx) for r in RUNS.values())
    if already:
        return

    loaded = load_runs_state(ctx)

    for run_id, run in loaded.items():
        run["scope_id"] = scope_id(ctx)
        run["company_id"] = ctx.company_id
        run["user_id"] = ctx.user_id
        RUNS[prefix + run_id] = run

    if _prune_run_retention_in_memory(ctx):
        save_runs_state(ctx)


def local_run_key(ctx: WorkerContext, run_id: str) -> str:
    prefix = f"{scope_id(ctx)}:"
    if run_id.startswith(prefix):
        return run_id
    return prefix + run_id


def public_run_id(ctx: WorkerContext, run_id: str) -> str:
    prefix = f"{scope_id(ctx)}:"
    return run_id.replace(prefix, "", 1)


def root_id_of(run: Dict[str, Any]) -> str:
    return run.get("root_id") or run.get("run_id")


def _remove_run_artifacts_for_retention(ctx: WorkerContext, root_id: str) -> None:
    base = member_runs_root(ctx)

    try:
        root_folder = safe_path_inside(base, os.path.join(base, root_id))
        if os.path.isdir(root_folder):
            shutil.rmtree(root_folder, ignore_errors=True)
    except Exception:
        logger.exception("Falha ao remover pasta expirada da run %s", root_id)

    try:
        zip_dir = safe_path_inside(base, os.path.join(base, "_zips"))
        if os.path.isdir(zip_dir):
            for zip_file in Path(zip_dir).glob(f"{root_id}*.zip"):
                if zip_file.is_file():
                    zip_file.unlink(missing_ok=True)
    except Exception:
        logger.exception("Falha ao remover ZIP expirado da run %s", root_id)


def _prune_run_retention_in_memory(ctx: WorkerContext) -> bool:
    scope = scope_id(ctx)
    now = now_ms()
    cutoff = now - max(1, RUN_RETENTION_DAYS) * 24 * 60 * 60 * 1000

    scope_runs = [
        (run_key, run)
        for run_key, run in RUNS.items()
        if run.get("scope_id") == scope
    ]

    active_roots = {
        root_id_of(run)
        for _, run in scope_runs
        if run.get("status") in ACTIVE_STATUSES
    }

    roots = [
        run
        for _, run in scope_runs
        if run.get("visible", True) and run.get("run_id") == root_id_of(run)
    ]
    roots.sort(key=lambda item: item.get("created_at", 0), reverse=True)

    delete_roots = {
        root_id_of(run)
        for run in roots
        if root_id_of(run) not in active_roots and int(run.get("created_at") or 0) < cutoff
    }

    if MAX_RUNS_PER_MEMBER > 0:
        kept = 0
        for run in roots:
            root_id = root_id_of(run)
            if root_id in delete_roots:
                continue
            kept += 1
            if root_id in active_roots:
                continue
            if kept > MAX_RUNS_PER_MEMBER:
                delete_roots.add(root_id)

    if not delete_roots:
        return False

    for run_key, run in list(scope_runs):
        if root_id_of(run) in delete_roots:
            RUNS.pop(run_key, None)

    for root_id in delete_roots:
        _remove_run_artifacts_for_retention(ctx, root_id)

    return True


def attempt_dir(ctx: WorkerContext, root_id: str, attempt_number: int) -> str:
    return os.path.join(member_runs_root(ctx), root_id, f"tentativa_{attempt_number}")


def flow_dir_for_task(run_dir: str, flow_mode: str) -> str:
    return os.path.join(run_dir, flow_mode)


def is_valid_nonempty_file(path: str) -> bool:
    return os.path.isfile(path)


def validate_output_integrity(path: str) -> None:
    if not path or not os.path.exists(path):
        raise RuntimeError(f"Saída não encontrada: {path}")


def list_run_files(run_dir: str, run_id: str) -> List[Dict[str, Any]]:
    files: List[Dict[str, Any]] = []

    if not os.path.exists(run_dir):
        return files

    for root, _, filenames in os.walk(run_dir):
        for filename in filenames:
            filename_low = filename.lower()

            if should_hide_run_file(filename):
                continue

            full_path = os.path.join(root, filename)

            if not is_valid_nonempty_file(full_path):
                continue

            rel = os.path.relpath(full_path, run_dir).replace("\\", "/")

            files.append(
                {
                    "run_id": run_id,
                    "name": filename,
                    "relative_path": rel,
                    "size": os.path.getsize(full_path),
                }
            )

    files.sort(key=lambda x: x["relative_path"].lower())
    return files


def files_for_result(all_files: List[Dict[str, Any]], result: Dict[str, Any]) -> List[Dict[str, Any]]:
    flow = result.get("flow_mode", "")
    cnpj = normalize_cnpj(result.get("cnpj", ""))

    if not flow or not cnpj:
        return []

    prefix = f"{flow}/{cnpj}"
    direct = [f for f in all_files if str(f.get("relative_path", "")).startswith(prefix)]

    if direct or flow != "notas":
        return direct

    codigo = str(result.get("codigo_dominio") or "").strip().lower()
    empresa = str(result.get("empresa") or result.get("nome_empresa") or "").strip().lower()
    return [
        item
        for item in all_files
        if str(item.get("relative_path", "")).lower().startswith("notas/")
        and (
            (codigo and codigo in str(item.get("relative_path", "")).lower())
            or (empresa and empresa in str(item.get("relative_path", "")).lower())
        )
    ]


def read_run_logs(run_dir: str, limit_chars: int = 80_000) -> str:
    log_file = os.path.join(run_dir, "logs.txt")

    if not os.path.exists(log_file):
        return ""

    with open(log_file, "r", encoding="utf-8", errors="replace") as f:
        data = f.read()

    if len(data) > limit_chars:
        return data[-limit_chars:]

    return data


def read_run_logs_filtered(
    run_dir: str,
    *,
    cnpj: str = "",
    flow: str = "",
    limit_chars: int = 80_000,
) -> str:
    log_file = os.path.join(run_dir, "logs.txt")

    if not os.path.exists(log_file):
        return ""

    cnpj_norm = normalize_cnpj(cnpj) if str(cnpj or "").strip() else ""
    flow_norm = str(flow or "").strip()
    max_chars = max(5_000, int(limit_chars or 80_000))
    lines: List[str] = []
    total_chars = 0

    with open(log_file, "r", encoding="utf-8", errors="replace") as f:
        for raw_line in f:
            line = raw_line.rstrip("\n")
            if cnpj_norm and cnpj_norm not in line:
                continue
            if flow_norm and f"flow={flow_norm}" not in line and f"flow_mode={flow_norm}" not in line:
                continue

            lines.append(line)
            total_chars += len(line) + 1

            while lines and total_chars > max_chars:
                removed = lines.pop(0)
                total_chars -= len(removed) + 1

    return "\n".join(lines)


def logs_by_attempt_for_root_filtered(
    ctx: WorkerContext,
    root_id: str,
    *,
    cnpj: str = "",
    flow: str = "",
    attempt_run_id: str = "",
    limit_chars: int = 80_000,
) -> List[Dict[str, Any]]:
    attempts = [run for run in runs_for_member(ctx) if root_id_of(run) == root_id]
    attempts.sort(key=lambda r: r.get("created_at", 0))

    output: List[Dict[str, Any]] = []
    wanted_attempt = str(attempt_run_id or "").strip()

    for run in attempts:
        if wanted_attempt and run.get("run_id") != wanted_attempt:
            continue

        run_dir = run.get("run_dir", "")
        logs = read_run_logs_filtered(
            run_dir,
            cnpj=cnpj,
            flow=flow,
            limit_chars=limit_chars,
        ) if run_dir else ""

        if not logs:
            continue

        output.append(
            {
                "run_id": run.get("run_id"),
                "attempt_number": run.get("attempt_number"),
                "attempt_type": run.get("attempt_type"),
                "status": run.get("status"),
                "logs": logs,
                "log_scope": "cnpj_flow",
            }
        )

    if wanted_attempt:
        return output

    return output[-1:] if output else []


def runs_for_member(ctx: WorkerContext) -> List[Dict[str, Any]]:
    ensure_member_runs_loaded(ctx)
    return [r for r in RUNS.values() if r.get("scope_id") == scope_id(ctx)]


def has_active_run(ctx: WorkerContext) -> bool:
    return any(run.get("status") in ACTIVE_STATUSES for run in runs_for_member(ctx))


def assert_no_active_run(ctx: WorkerContext) -> None:
    if has_active_run(ctx):
        raise HTTPException(
            status_code=409,
            detail="Você já tem uma run em execução ou na fila. Aguarde terminar antes de criar, duplicar ou dar retry.",
        )


def attempts_for_root(ctx: WorkerContext, root_id: str) -> List[Dict[str, Any]]:
    attempts = [
        compact_attempt(ctx, run, include_logs=False)
        for run in runs_for_member(ctx)
        if root_id_of(run) == root_id
    ]
    attempts.sort(key=lambda r: r.get("created_at", 0))
    return attempts


def logs_by_attempt_for_root(ctx: WorkerContext, root_id: str, limit_chars: int = 80_000) -> List[Dict[str, Any]]:
    attempts = [run for run in runs_for_member(ctx) if root_id_of(run) == root_id]
    attempts.sort(key=lambda r: r.get("created_at", 0))

    output: List[Dict[str, Any]] = []

    for run in attempts:
        run_dir = run.get("run_dir", "")
        output.append(
            {
                "run_id": run.get("run_id"),
                "attempt_number": run.get("attempt_number"),
                "attempt_type": run.get("attempt_type"),
                "status": run.get("status"),
                "logs": read_run_logs(run_dir, limit_chars=limit_chars) if run_dir else "",
            }
        )

    return output


def collect_files_for_root(ctx: WorkerContext, root_id: str) -> List[Dict[str, Any]]:
    output: List[Dict[str, Any]] = []
    attempts = [run for run in runs_for_member(ctx) if root_id_of(run) == root_id]
    attempts.sort(key=lambda r: r.get("created_at", 0))

    for run in attempts:
        run_dir = run.get("run_dir", "")
        run_id = run.get("run_id", "")

        if run_dir:
            output.extend(list_run_files(run_dir, run_id))

    return output


def aggregate_results_for_root(ctx: WorkerContext, root_id: str) -> List[Dict[str, Any]]:
    attempts = [run for run in runs_for_member(ctx) if root_id_of(run) == root_id]
    attempts.sort(key=lambda r: r.get("created_at", 0))

    by_key: Dict[str, Dict[str, Any]] = {}

    for attempt in attempts:
        attempt_number = attempt.get("attempt_number")
        attempt_status = attempt.get("status")

        if attempt_status in ACTIVE_STATUSES:
            finished_or_known_keys = {task_key(r) for r in attempt.get("results", [])}

            for task in attempt.get("input_tasks", []):
                key = task_key(task)

                if key not in finished_or_known_keys:
                    pending = dict(task)
                    pending.update(
                        {
                            "status": "queued",
                            "retryable": False,
                            "attempt_number": attempt_number,
                            "source_run_id": attempt.get("run_id"),
                            "source_attempt_number": attempt_number,
                        }
                    )
                    by_key[key] = pending

        for result in attempt.get("results", []):
            key = task_key(result)
            merged = dict(result)
            merged["attempt_number"] = attempt_number
            merged["source_run_id"] = attempt.get("run_id")
            merged["source_attempt_number"] = attempt_number
            by_key[key] = merged

    root = next((r for r in runs_for_member(ctx) if r.get("run_id") == root_id), None)

    if root:
        ordered_keys = [task_key(item) for item in root.get("input_tasks", [])]
        ordered_set = set(ordered_keys)
        ordered = [by_key[k] for k in ordered_keys if k in by_key]
        rest = [value for key, value in by_key.items() if key not in ordered_set]
        return ordered + rest

    return list(by_key.values())


def recompute_run_counters(run_key: str) -> None:
    run = RUNS.get(run_key)

    if not run:
        return

    results = run.get("results", [])

    run["done"] = len([r for r in results if r.get("status") in ITEM_FINAL_STATUSES])
    run["ok"] = len([r for r in results if r.get("status") == "ok"])
    run["erros"] = len([r for r in results if r.get("status") == "erro"])
    run["running"] = len([r for r in results if r.get("status") == "running"])


async def upsert_run_result(ctx: WorkerContext, run_key: str, result: Dict[str, Any], *, save: bool = False) -> None:
    async with RUN_LOCK:
        run = RUNS.get(run_key)

        if not run:
            return

        results = list(run.get("results", []))
        key = result_merge_key(result)
        replaced = False

        for idx, old in enumerate(results):
            if result_merge_key(old) == key:
                results[idx] = result
                replaced = True
                break

        if not replaced:
            results.append(result)

        run["results"] = results
        run["files"] = list_run_files(run.get("run_dir", ""), run.get("run_id", ""))
        recompute_run_counters(run_key)

        if save:
            save_runs_state(ctx)


def write_run_log(run_log_file: str, line: str) -> None:
    try:
        Path(os.path.dirname(run_log_file)).mkdir(parents=True, exist_ok=True)

        with open(run_log_file, "a", encoding="utf-8") as f:
            f.write(line.rstrip() + "\n")
    except Exception:
        pass


def item_stop_requested(run_key: str, item_or_result: Dict[str, Any]) -> bool:
    run = RUNS.get(run_key)

    if not run:
        return False

    if bool(run.get("stop_requested")):
        return True

    key = task_key(item_or_result)
    return key in set(run.get("stop_requested_keys", []) or [])


def make_cancelled_result(item: Dict[str, Any], *, reason: str, code: str) -> Dict[str, Any]:
    flow_mode = item.get("flow_mode", "")

    return {
        "cnpj": item.get("cnpj", ""),
        "codigo_dominio": item.get("codigo_dominio", ""),
        "nome_empresa": item.get("nome_empresa", ""),
        "account_id": item.get("account_id", ""),
        "account_alias": item.get("account_alias", ""),
        "flow_mode": flow_mode,
        "flow_label": FLOW_LABELS.get(flow_mode, flow_mode),
        "status": "cancelled",
        "erro": reason,
        "erro_code": code,
        "erro_action": "Executar retry se este fluxo ainda for necessário.",
        "retryable": True,
        "finished_at": now_ms(),
    }


def make_interrupted_result(item: Dict[str, Any], *, reason: str, code: str) -> Dict[str, Any]:
    result = make_cancelled_result(item, reason=reason, code=code)
    result["status"] = "interrompida"
    result["erro_action"] = "Executar retry se este fluxo ainda for necessário."
    return result


async def cancel_item_before_start(
    ctx: WorkerContext,
    run_key: str,
    item: Dict[str, Any],
    run_log_file: str,
    *,
    reason: str,
    code: str,
) -> Dict[str, Any]:
    result = make_cancelled_result(item, reason=reason, code=code)

    write_run_log(
        run_log_file,
        (
            f"[ITEM_CANCELLED] flow={item.get('flow_mode', '')} "
            f"cnpj={normalize_cnpj(item.get('cnpj', ''))} "
            f"conta={item.get('account_alias', '')} code={code} msg={reason}"
        ),
    )

    await upsert_run_result(ctx, run_key, result, save=True)
    return result


def _result_is_active_running(result: Optional[Dict[str, Any]]) -> bool:
    return bool(result and result.get("status") == "running")


def _result_is_final(result: Optional[Dict[str, Any]]) -> bool:
    return bool(result and result.get("status") in ITEM_FINAL_STATUSES)


def _find_result(results: List[Dict[str, Any]], key: str) -> Optional[Dict[str, Any]]:
    for result in results:
        if task_key(result) == key:
            return result
    return None


def _upsert_result_in_list(results: List[Dict[str, Any]], result: Dict[str, Any]) -> List[Dict[str, Any]]:
    key = task_key(result)
    output = list(results)

    for idx, old in enumerate(output):
        if task_key(old) == key:
            output[idx] = result
            return output

    output.append(result)
    return output


async def request_stop_for_attempt(
    ctx: WorkerContext,
    attempt_run_key: str,
    *,
    only_key: Optional[str] = None,
    reason: str,
    code: str,
) -> Dict[str, Any]:
    async with RUN_LOCK:
        run = RUNS.get(attempt_run_key)

        if not run or run.get("scope_id") != scope_id(ctx):
            raise HTTPException(status_code=404, detail="Tentativa da run não encontrada.")

        if run.get("status") not in ACTIVE_STATUSES:
            return {
                "changed": False,
                "message": "A tentativa já terminou.",
                "attempt_run_id": run.get("run_id"),
            }

        already_requested = bool(run.get("stop_requested")) if only_key is None else only_key in set(run.get("stop_requested_keys", []) or [])

        if only_key is None:
            run["stop_requested"] = True
        else:
            keys = set(run.get("stop_requested_keys", []) or [])
            keys.add(only_key)
            run["stop_requested_keys"] = sorted(keys)

        results = list(run.get("results", []))
        cancelled = 0
        interrupted_running = 0
        interrupted_running_flows = set()
        run_log_file = run.get("run_log_file") or os.path.join(run.get("run_dir", ""), "logs.txt")

        for item in run.get("input_tasks", []):
            key = task_key(item)

            if only_key is not None and key != only_key:
                continue

            existing = _find_result(results, key)

            if _result_is_active_running(existing):
                interrupted_result = make_interrupted_result(item, reason=reason, code=code)
                results = _upsert_result_in_list(results, interrupted_result)
                interrupted_running_flows.add(item.get("flow_mode", ""))
                write_run_log(
                    run_log_file,
                    (
                        f"[ITEM_INTERRUPTED] flow={item.get('flow_mode', '')} "
                        f"cnpj={normalize_cnpj(item.get('cnpj', ''))} "
                        f"conta={item.get('account_alias', '')} code={code} retryable=True msg={reason}"
                    ),
                )
                interrupted_running += 1
                continue

            if _result_is_final(existing):
                continue

            cancelled_result = make_cancelled_result(item, reason=reason, code=code)
            results = _upsert_result_in_list(results, cancelled_result)
            write_run_log(
                run_log_file,
                (
                    f"[ITEM_CANCELLED] flow={item.get('flow_mode', '')} "
                    f"cnpj={normalize_cnpj(item.get('cnpj', ''))} "
                    f"conta={item.get('account_alias', '')} code={code} msg={reason}"
                ),
            )
            cancelled += 1

        run["results"] = results
        run["files"] = list_run_files(run.get("run_dir", ""), run.get("run_id", ""))
        recompute_run_counters(attempt_run_key)

        if only_key is None:
            expected_keys = {task_key(item) for item in run.get("input_tasks", [])}
            final_keys = {task_key(result) for result in results if result.get("status") in ITEM_FINAL_STATUSES}
            can_finish_now = expected_keys.issubset(final_keys) and "escrituracao" not in interrupted_running_flows
            if can_finish_now:
                run["status"] = "cancelled"
                run["finished_at"] = run.get("finished_at") or now_ms()
                run["active_groups"] = []

        save_runs_state(ctx)

        return {
            "changed": True,
            "already_requested": already_requested,
            "message": "A parada já tinha sido solicitada e a execução ainda está finalizando." if already_requested else "Parada solicitada.",
            "attempt_run_id": run.get("run_id"),
            "cancelled_pending": cancelled,
            "interrupted_running": interrupted_running,
        }


def compact_attempt(ctx: WorkerContext, run: Dict[str, Any], *, include_logs: bool) -> Dict[str, Any]:
    data = sanitize_run_for_response(run)
    run_dir = data.get("run_dir")
    run_id = data.get("run_id")

    if run_dir:
        data["files"] = list_run_files(run_dir, run_id)
        if include_logs:
            data["logs"] = read_run_logs(run_dir)

    return data


def compact_root_run(
    ctx: WorkerContext,
    run: Dict[str, Any],
    *,
    include_logs: bool,
    include_attempts: bool,
    include_logs_by_attempt: bool,
    include_result_files: bool = True,
) -> Dict[str, Any]:
    root_id = root_id_of(run)
    attempts = [r for r in runs_for_member(ctx) if root_id_of(r) == root_id]
    attempts.sort(key=lambda r: r.get("created_at", 0))

    active_attempts = [r for r in attempts if r.get("status") in ACTIVE_STATUSES]
    any_running = bool(active_attempts)
    any_stop_requested = any(bool(r.get("stop_requested")) for r in active_attempts)
    latest = attempts[-1] if attempts else run

    results = aggregate_results_for_root(ctx, root_id)
    files = collect_files_for_root(ctx, root_id)

    if include_result_files:
        results = [{**r, "files": files_for_result(files, r)} for r in results]

    ok_count = len([r for r in results if r.get("status") == "ok"])
    err_count = len([r for r in results if r.get("status") == "erro"])
    cancelled_count = len([r for r in results if r.get("status") == "cancelled"])
    interrupted_count = len([r for r in results if r.get("status") == "interrompida"])
    running_count = len([r for r in results if r.get("status") == "running"])
    total = len(run.get("input_tasks", [])) or run.get("total", 0)

    if any_running:
        done = min(len([r for r in results if r.get("status") in ITEM_FINAL_STATUSES]), total) if total else 0
        status = "running" if running_count else "queued"
    else:
        status = latest.get("status", run.get("status"))
        if status in FINAL_STATUSES:
            done = total
        else:
            done = min(len([r for r in results if r.get("status") in ITEM_FINAL_STATUSES]), total) if total else len(results)

    flow_counts = {flow: 0 for flow in FLOW_ORDER}

    for task in run.get("input_tasks", []):
        flow = task.get("flow_mode")
        if flow in flow_counts:
            flow_counts[flow] += 1

    data = sanitize_run_for_response(run)

    data.update(
        {
            "status": status,
            "total": total,
            "done": done,
            "ok": ok_count,
            "erros": err_count,
            "cancelled": cancelled_count,
            "interrompidas": interrupted_count,
            "running": running_count,
            "results": results,
            "files": files,
            "attempts_count": len(attempts),
            "latest_attempt_run_id": latest.get("run_id"),
            "latest_attempt_number": latest.get("attempt_number"),
            "flow_counts": flow_counts,
            "has_active_run": has_active_run(ctx),
            "stop_requested": any_stop_requested,
            "stopping": any_running and any_stop_requested,
            "usar_codigo_dominio": bool(run.get("usar_codigo_dominio", True)),
            "reabrir_escrituracao_fechada": bool(run.get("reabrir_escrituracao_fechada", True)),
        }
    )

    if include_attempts:
        data["attempts"] = attempts_for_root(ctx, root_id)

    if include_logs_by_attempt:
        data["logs_by_attempt"] = logs_by_attempt_for_root(ctx, root_id)

    if include_logs:
        data["logs"] = "\n\n".join(read_run_logs(r.get("run_dir", "")) for r in attempts if r.get("run_dir"))

    return data


def visible_root_runs(ctx: WorkerContext) -> List[Dict[str, Any]]:
    ensure_member_runs_loaded(ctx)
    if _prune_run_retention_in_memory(ctx):
        save_runs_state(ctx)

    roots = [
        run
        for run in RUNS.values()
        if run.get("visible", True) and run.get("run_id") == root_id_of(run)
        and run.get("scope_id") == scope_id(ctx)
    ]

    data = [
        compact_root_run(
            ctx,
            run,
            include_logs=False,
            include_attempts=False,
            include_logs_by_attempt=False,
            include_result_files=False,
        )
        for run in roots
    ]

    data.sort(key=lambda r: r.get("created_at", 0), reverse=True)
    return data


def hydrate_tasks_with_current_accounts(ctx: WorkerContext, tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    hydrated: List[Dict[str, Any]] = []

    for item in tasks:
        account_id = item.get("account_id", "")
        account = get_account_or_404(ctx, account_id)
        enriched = dict(item)
        enriched["account_alias"] = account.get("alias", "")
        enriched["usuario"] = account.get("usuario", "")
        enriched["senha"] = account.get("senha", "")
        enriched["usar_codigo_dominio"] = bool(enriched.get("usar_codigo_dominio", True))
        enriched["reabrir_escrituracao_fechada"] = bool(enriched.get("reabrir_escrituracao_fechada", True))
        hydrated.append(enriched)

    return hydrated


def hydrate_tasks_with_retry_snapshot(
    ctx: WorkerContext,
    root_run: Dict[str, Any],
    tasks: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    snapshots = root_run.get("credential_snapshots", {}) or {}
    hydrated: List[Dict[str, Any]] = []

    for item in tasks:
        key = task_key(item)
        snap = snapshots.get(key)
        enriched = dict(item)
        enriched["usar_codigo_dominio"] = bool(enriched.get("usar_codigo_dominio", root_run.get("usar_codigo_dominio", True)))
        enriched["reabrir_escrituracao_fechada"] = bool(
            enriched.get("reabrir_escrituracao_fechada", root_run.get("reabrir_escrituracao_fechada", True))
        )

        if snap:
            enriched["account_alias"] = snap.get("account_alias", enriched.get("account_alias", ""))
            enriched["usuario"] = snap.get("usuario", "")
            enriched["senha"] = snap.get("senha", "")
        else:
            account = get_account_or_404(ctx, enriched.get("account_id", ""))
            enriched["account_alias"] = account.get("alias", "")
            enriched["usuario"] = account.get("usuario", "")
            enriched["senha"] = account.get("senha", "")

        hydrated.append(enriched)

    return hydrated


def build_tasks_from_dataset(
    ctx: WorkerContext,
    dataset: Dict[str, Any],
    flow_selection: Dict[str, Dict[str, bool]],
    *,
    usar_codigo_dominio: bool = True,
    reabrir_escrituracao_fechada: bool = True,
) -> List[Dict[str, Any]]:
    validation = validate_dataset_items(ctx, dataset.get("items", []))

    if not validation["valid"]:
        raise HTTPException(
            status_code=400,
            detail={"message": "Conjunto possui erros. Corrija antes de executar.", "validation": validation},
        )

    items = normalize_valid_dataset_items(validation["items"])
    tasks: List[Dict[str, Any]] = []

    for item in items:
        cnpj_digits = normalize_cnpj(item.get("cnpj_digits") or item.get("cnpj", ""))

        flows = normalize_flows(
            flow_selection.get(cnpj_digits)
            or flow_selection.get(item.get("cnpj", ""))
            or default_flows()
        )

        ordered_flows = [flow for flow in FLOW_EXECUTION_ORDER if flows.get(flow)]

        for flow in ordered_flows:
            tasks.append(
                {
                    "cnpj": item.get("cnpj", ""),
                    "cnpj_digits": cnpj_digits,
                    "codigo_dominio": item.get("codigo_dominio", ""),
                    "nome_empresa": item.get("nome_empresa", ""),
                    "account_id": item.get("account_id", ""),
                    "flow_mode": flow,
                    "flow_label": FLOW_LABELS.get(flow, flow),
                    "usar_codigo_dominio": bool(usar_codigo_dominio and item.get("codigo_dominio")) if flow == "notas" else bool(usar_codigo_dominio),
                    "reabrir_escrituracao_fechada": bool(reabrir_escrituracao_fechada),
                }
            )

    return hydrate_tasks_with_current_accounts(ctx, tasks)


def group_tasks_by_cnpj(runtime_tasks: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    groups: Dict[str, List[Dict[str, Any]]] = {}

    for item in runtime_tasks:
        key = normalize_cnpj(item.get("cnpj_digits") or item.get("cnpj", ""))
        groups.setdefault(key, []).append(item)

    return groups


def root_attempt_ids(ctx: WorkerContext, root_id: str) -> List[str]:
    return [
        rid
        for rid, run in RUNS.items()
        if run.get("scope_id") == scope_id(ctx) and root_id_of(run) == root_id
    ]


def attempts_for_root_raw(ctx: WorkerContext, root_id: str) -> List[Dict[str, Any]]:
    attempts = [run for run in runs_for_member(ctx) if root_id_of(run) == root_id]
    attempts.sort(key=lambda r: r.get("created_at", 0))
    return attempts


def root_has_active_attempt(ctx: WorkerContext, root_id: str) -> bool:
    return any(
        run.get("status") in ACTIVE_STATUSES
        for run in runs_for_member(ctx)
        if root_id_of(run) == root_id
    )


def assert_root_not_active(ctx: WorkerContext, root_id: str) -> None:
    if root_has_active_attempt(ctx, root_id):
        raise HTTPException(
            status_code=409,
            detail="Não é possível executar esta ação com a run em execução ou na fila. Pare a run primeiro e aguarde finalizar/cancelar.",
        )


def delete_run_folder(ctx: WorkerContext, root_id: str) -> Dict[str, Any]:
    root_folder = safe_path_inside(member_runs_root(ctx), os.path.join(member_runs_root(ctx), root_id))

    removed_folder = False
    removed_zip = False

    if os.path.exists(root_folder):
        if not os.path.isdir(root_folder):
            raise HTTPException(status_code=400, detail="Caminho da run não é uma pasta.")

        shutil.rmtree(root_folder)
        removed_folder = True

    zip_path = safe_path_inside(
        member_runs_root(ctx),
        os.path.join(member_runs_root(ctx), "_zips", f"{root_id}.zip"),
    )

    if os.path.exists(zip_path):
        if os.path.isfile(zip_path):
            os.remove(zip_path)
            removed_zip = True

    return {
        "removed_folder": removed_folder,
        "removed_zip": removed_zip,
        "folder": root_folder,
        "zip": zip_path,
    }


def delete_run_from_memory_and_db(ctx: WorkerContext, root_id: str) -> Dict[str, Any]:
    attempt_ids = root_attempt_ids(ctx, root_id)

    if not attempt_ids:
        raise HTTPException(status_code=404, detail="Run não encontrada.")

    for attempt_id in attempt_ids:
        RUNS.pop(attempt_id, None)

    save_runs_state(ctx)

    return {
        "removed_attempt_ids": [public_run_id(ctx, item) for item in attempt_ids],
        "removed_attempts": len(attempt_ids),
    }


def should_hide_run_file(filename: str) -> bool:
    filename_low = filename.lower()

    if filename_low.endswith(".zip"):
        return True

    if filename_low in {"logs.txt", "summary.json", "cnpjs_com_erro.txt"}:
        return True

    if filename_low.startswith("erro_") and filename_low.endswith(".png"):
        return True

    return False


def should_skip_zip_file(filename: str) -> bool:
    return should_hide_run_file(filename)


def collect_unified_zip_entries(ctx: WorkerContext, root_id: str, cnpj: str = "") -> Dict[str, str]:
    entries: Dict[str, str] = {}
    attempts = attempts_for_root_raw(ctx, root_id)
    cnpj_digits = normalize_cnpj(cnpj) if str(cnpj or "").strip() else ""

    if not attempts:
        raise HTTPException(status_code=404, detail="Run não encontrada.")

    for attempt in attempts:
        attempt_dir_path = attempt.get("run_dir", "")

        if not attempt_dir_path or not os.path.isdir(attempt_dir_path):
            continue

        attempt_dir_path = safe_path_inside(member_runs_root(ctx), attempt_dir_path)

        all_files = list_run_files(attempt_dir_path, attempt.get("run_id", ""))
        successful_results = [
            result
            for result in attempt.get("results", [])
            if result.get("status") == "ok"
            and (not cnpj_digits or normalize_cnpj(result.get("cnpj", "")) == cnpj_digits)
        ]

        for result in successful_results:
            for item in files_for_result(all_files, result):
                rel = str(item.get("relative_path", ""))
                full_path = safe_path_inside(attempt_dir_path, os.path.join(attempt_dir_path, rel))
                if rel and is_valid_nonempty_file(full_path):
                    entries[rel] = full_path

    return entries


def create_root_zip(ctx: WorkerContext, root_id: str) -> str:
    root_folder = os.path.join(member_runs_root(ctx), root_id)

    if not os.path.isdir(root_folder):
        raise HTTPException(status_code=404, detail="Pasta da run não encontrada.")

    if root_has_active_attempt(ctx, root_id):
        raise HTTPException(status_code=409, detail="Não é possível baixar o ZIP enquanto a run está em execução ou na fila.")

    zip_dir = os.path.join(member_runs_root(ctx), "_zips")
    Path(zip_dir).mkdir(parents=True, exist_ok=True)
    zip_path = os.path.join(zip_dir, f"{root_id}.zip")

    if os.path.exists(zip_path):
        os.remove(zip_path)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        entries = collect_unified_zip_entries(ctx, root_id)
        if not entries:
            zf.writestr(
                "SEM_ARQUIVOS_GERADOS.txt",
                (
                    "Esta run foi concluída, mas não gerou arquivos baixáveis.\n"
                    "Isso pode acontecer quando o fluxo não encontrou notas, DAMs ou certidões para baixar.\n"
                    "Consulte o painel da run e os logs para ver o resultado de cada empresa.\n"
                ),
            )
        for rel, full_path in sorted(entries.items(), key=lambda x: x[0].lower()):
            zf.write(full_path, rel)

    return zip_path


def _zip_arcname_for_cnpj(rel: str, used: set) -> str:
    parts = str(rel or "").replace("\\", "/").split("/")
    flow = parts[0] if parts and parts[0] in FLOW_ORDER else "arquivos"
    filename = os.path.basename(rel) or "arquivo"
    arcname = f"{flow}/{filename}"

    if arcname not in used:
        used.add(arcname)
        return arcname

    stem, ext = os.path.splitext(filename)
    idx = 2
    while True:
        candidate = f"{flow}/{stem}_{idx}{ext}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        idx += 1


def create_cnpj_zip(ctx: WorkerContext, root_id: str, cnpj: str) -> str:
    cnpj_digits = normalize_cnpj(cnpj)

    if len(cnpj_digits) != 14:
        raise HTTPException(status_code=400, detail="CNPJ inválido.")

    root_folder = os.path.join(member_runs_root(ctx), root_id)

    if not os.path.isdir(root_folder):
        raise HTTPException(status_code=404, detail="Pasta da run não encontrada.")

    if root_has_active_attempt(ctx, root_id):
        raise HTTPException(status_code=409, detail="Não é possível baixar o ZIP enquanto a run está em execução ou na fila.")

    entries = collect_unified_zip_entries(ctx, root_id, cnpj_digits)

    if not entries:
        raise HTTPException(status_code=404, detail="Nenhum arquivo encontrado para este CNPJ.")

    zip_dir = os.path.join(member_runs_root(ctx), "_zips")
    Path(zip_dir).mkdir(parents=True, exist_ok=True)
    zip_path = os.path.join(zip_dir, f"{root_id}_{cnpj_digits}.zip")

    if os.path.exists(zip_path):
        os.remove(zip_path)

    used: set = set()

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for rel, full_path in sorted(entries.items(), key=lambda x: x[0].lower()):
            zf.write(full_path, _zip_arcname_for_cnpj(rel, used))

    return zip_path


def build_state_sync(ctx: WorkerContext) -> Dict[str, Any]:
    ensure_member_runs_loaded(ctx)
    usage = account_usage(ctx)

    return {
        "ts": now_ms(),
        "worker_public_url": WORKER_PUBLIC_URL,
        "via_worker": ctx.via_worker,
        "company": {
            "id": ctx.company_id,
            "name": ctx.company_name,
        },
        "colaborador": {
            "id": ctx.user_id,
            "email": ctx.user_email,
            "role": ctx.user_role,
        },
        "scope": {
            "id": scope_id(ctx),
            "storage": "member",
        },
        "has_active_run": has_active_run(ctx),
        "accounts": [
            {**acc, "linked_cnpjs": usage.get(acc["id"], [])}
            for acc in load_accounts_public(ctx)
        ],
        "datasets": [compact_dataset_record(ds) for ds in load_datasets(ctx)],
        "runs": visible_root_runs(ctx),
        "config": {
            "max_browsers": MAX_BROWSERS,
            "base_browsers": BASE_BROWSER_SLOTS,
            "browser_turbo_extra": BROWSER_TURBO_EXTRA,
            "browser_pool_configured": BROWSER_POOL_CONFIGURED,
            "headless": HEADLESS,
            "max_datasets": None,
            "max_runs_per_member": MAX_RUNS_PER_MEMBER,
            "run_retention_days": RUN_RETENTION_DAYS,
            "flows": [{"id": flow, "label": FLOW_LABELS[flow]} for flow in FLOW_ORDER],
            "flow_execution_order": FLOW_EXECUTION_ORDER,
            "dam_requires_escrituracao": False,
            "dam_runs_after_own_escrituracao_only": True,
            "dam_without_escrituracao_allowed": True,
            "notas_layout_modes": [
                {
                    "id": "codigo_dominio",
                    "label": "Código Domínio",
                    "usar_codigo_dominio": True,
                },
                {
                    "id": "cnpj_empresa",
                    "label": "CNPJ + Empresa",
                    "usar_codigo_dominio": False,
                },
            ],
            "retry_options": {
                "include_cancelled": True,
                "only_retryable": False,
            },
            "zip_unificado_sem_tentativas": True,
            "zip_blocked_while_running": True,
            "queue_mode": "global_fair_round_robin_by_member",
        },
    }
